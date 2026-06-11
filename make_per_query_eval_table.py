#!/usr/bin/env python3
import argparse
import csv
import json
import re
from pathlib import Path
from collections import defaultdict, Counter

MODELS = ["EVIQUE", "VideoRAG", "NaiveRAG", "TextVideoRAG", "LightRAG", "GraphRAG-l", "GraphRAG-g"]
METRICS = ["Comprehensiveness", "Empowerment", "Trustworthiness", "Depth", "Density", "Overall Score"]
WIN_METRICS = METRICS[:-1] + ["Overall Winner"]

QUERY_FIELDS = ("query_id", "qid", "id", "uid")
LEFT_MODEL_FIELDS = ("left_model", "model_a", "model_1", "model1")
RIGHT_MODEL_FIELDS = ("right_model", "model_b", "model_2", "model2")
COMPARISON_FIELDS = ("comparison", "pair", "model_pair")
RUN_FIELDS = ("run_id", "eval_run", "run", "trial")
DIRECTION_FIELDS = ("direction", "order", "pair_order", "is_reversed", "reverse", "reversed", "swapped")
ANSWER_1_FIELDS = ("answer_1_model", "answer1_model", "answer_a_model", "first_model")
ANSWER_2_FIELDS = ("answer_2_model", "answer2_model", "answer_b_model", "second_model")
WINNER_FIELDS = ("winner", "winner_model", "selected", "choice", "verdict", "overall_winner", "Winner")
EXPLANATION_FIELDS = ("explanation", "Explanation", "reason", "rationale")

def natural_qid_key(qid):
    m = re.match(r"([a-zA-Z]+)(\d+)", str(qid))
    return (m.group(1), int(m.group(2))) if m else (str(qid), 0)

def load_json(path):
    with Path(path).open("r", encoding="utf-8") as f:
        return json.load(f)

def load_questions(path):
    out = {}
    if not path or not Path(path).exists():
        return out
    raw = load_json(path)
    if isinstance(raw, list):
        for item in raw:
            if isinstance(item, dict):
                qid = str(item.get("uid") or item.get("id") or item.get("query_id") or "")
                q = item.get("question") or item.get("query") or ""
                if qid:
                    out[qid] = q
    return out

def write_csv(path, rows):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    keys = []
    for row in rows:
        for k in row:
            if k not in keys:
                keys.append(k)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

def parse_quantitative(qraw, qmap):
    rows = []

    for key, val in qraw.items():
        # expected key: v1::EVIQUE::run1
        parts = key.split("::")
        if len(parts) < 3:
            continue

        qid = parts[0]
        model = parts[1]
        run = parts[2]

        if not isinstance(val, dict):
            continue

        for metric in METRICS:
            item = val.get(metric)
            if not isinstance(item, dict):
                continue
            score = item.get("Score")
            explanation = item.get("Explanation", "")
            if score is None:
                continue

            rows.append({
                "query_id": qid,
                "question": qmap.get(qid, ""),
                "model": model,
                "run": run,
                "metric": metric,
                "score": score,
                "explanation": explanation,
                "source_key": key,
            })

    rows.sort(key=lambda r: (
        natural_qid_key(r["query_id"]),
        MODELS.index(r["model"]) if r["model"] in MODELS else 99,
        r["run"],
        METRICS.index(r["metric"]) if r["metric"] in METRICS else 99,
    ))
    return rows

def quantitative_wide(q_long):
    bucket = defaultdict(list)
    qmap = {}

    for r in q_long:
        qid = r["query_id"]
        qmap[qid] = r.get("question", "")
        bucket[(qid, r["model"], r["metric"])].append(float(r["score"]))

    rows = []
    for qid in sorted(qmap, key=natural_qid_key):
        row = {
            "query_id": qid,
            "question": qmap[qid],
        }
        for model in MODELS:
            for metric in METRICS:
                vals = bucket.get((qid, model, metric), [])
                row[f"{model}_{metric}"] = round(sum(vals) / len(vals), 3) if vals else ""
        rows.append(row)
    return rows

def normalize_field_name(name):
    return re.sub(r"[\s_\-]+", "", str(name)).lower()

def available_columns(row):
    return sorted(str(k) for k in row.keys()) if isinstance(row, dict) else []

def get_field(row, names, default=""):
    if not isinstance(row, dict):
        return default
    by_norm = {normalize_field_name(k): k for k in row.keys()}
    for name in names:
        key = by_norm.get(normalize_field_name(name))
        if key is None:
            continue
        value = row.get(key)
        if value is not None and str(value).strip() != "":
            return value
    return default

def get_metric_item(row, metric):
    if not isinstance(row, dict):
        return None
    if metric in row:
        return row[metric]
    wanted = normalize_field_name(metric)
    for key, value in row.items():
        if normalize_field_name(key) == wanted:
            return value
    return None

def parse_comparison(value):
    if value is None:
        return None, None
    m = re.match(r"\s*(.+?)\s+vs\.?\s+(.+?)\s*$", str(value), flags=re.IGNORECASE)
    if not m:
        return None, None
    return m.group(1).strip(), m.group(2).strip()

def parse_winrate_source_key(source_key):
    parts = str(source_key).split("::")
    if len(parts) < 4 or parts[2].lower() != "vs":
        return {}
    meta = {
        "query_id": parts[0],
        "left_model": parts[1],
        "right_model": parts[3],
        "comparison": f"{parts[1]} vs {parts[3]}",
    }
    if len(parts) >= 5:
        meta["run_id"] = parts[4]
    if len(parts) >= 6:
        meta["direction"] = parts[5]
    return meta

def normalize_boolish(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)

    text = str(value).strip().lower()
    compact = re.sub(r"[^a-z0-9]+", "", text)
    if compact in {"rev", "reverse", "reversed", "swapped", "swap", "backward", "ba", "btoa", "rightleft", "1", "true", "yes"}:
        return True
    if compact in {"ori", "orig", "original", "forward", "normal", "direct", "ab", "atob", "leftright", "0", "false", "no"}:
        return False
    return None

def normalize_model_text(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())

def canonical_model(value, candidates=()):
    raw = str(value).strip()
    if raw == "":
        return ""
    known = []
    for model in list(candidates) + MODELS:
        if model and model not in known:
            known.append(model)
    raw_norm = normalize_model_text(raw)
    for model in known:
        if normalize_model_text(model) == raw_norm:
            return model
    return raw

def detect_reverse(row, meta):
    direction_value = get_field(row, DIRECTION_FIELDS, "")
    if direction_value == "" and meta:
        direction_value = meta.get("direction", "")

    if direction_value != "":
        parsed = normalize_boolish(direction_value)
        if parsed is not None:
            return parsed, str(direction_value).strip()

    return None, str(direction_value).strip()

def resolve_comparison_models(row, meta, source_label):
    qid = str(get_field(row, QUERY_FIELDS, meta.get("query_id", ""))).strip()
    comparison = str(get_field(row, COMPARISON_FIELDS, meta.get("comparison", ""))).strip()
    left_model = str(get_field(row, LEFT_MODEL_FIELDS, meta.get("left_model", ""))).strip()
    right_model = str(get_field(row, RIGHT_MODEL_FIELDS, meta.get("right_model", ""))).strip()

    parsed_left, parsed_right = parse_comparison(comparison)
    if not left_model and parsed_left:
        left_model = parsed_left
    if not right_model and parsed_right:
        right_model = parsed_right
    if not comparison and left_model and right_model:
        comparison = f"{left_model} vs {right_model}"

    missing = []
    if not qid:
        missing.append("query_id/qid/id")
    if not left_model:
        missing.append("left_model/model_a or parseable comparison")
    if not right_model:
        missing.append("right_model/model_b or parseable comparison")
    if missing:
        raise ValueError(
            f"Cannot identify winrate fields ({', '.join(missing)}) in {source_label}. "
            f"Available columns: {available_columns(row)}"
        )
    return qid, comparison, left_model, right_model

def resolve_answer_models(row, meta, left_model, right_model, source_label):
    answer_1_model = str(get_field(row, ANSWER_1_FIELDS, "")).strip()
    answer_2_model = str(get_field(row, ANSWER_2_FIELDS, "")).strip()
    is_reversed, direction = detect_reverse(row, meta)

    if answer_1_model and answer_2_model:
        answer_1_model = canonical_model(answer_1_model, (left_model, right_model))
        answer_2_model = canonical_model(answer_2_model, (left_model, right_model))
        if is_reversed is None:
            is_reversed = (
                normalize_model_text(answer_1_model) == normalize_model_text(right_model)
                and normalize_model_text(answer_2_model) == normalize_model_text(left_model)
            )
            direction = "rev" if is_reversed else "ori"
        return answer_1_model, answer_2_model, direction, is_reversed

    if is_reversed is None:
        raise ValueError(
            f"Cannot identify winrate direction/order in {source_label}; positional winners cannot be mapped safely. "
            f"Expected one of {list(DIRECTION_FIELDS)} or explicit answer_1_model/answer_2_model. "
            f"Available columns: {available_columns(row)}"
        )

    if is_reversed:
        return right_model, left_model, direction or "rev", True
    return left_model, right_model, direction or "ori", False

def normalize_winner(raw_winner, answer_1_model, answer_2_model, candidates):
    if raw_winner is None:
        return ""

    raw = str(raw_winner).strip()
    if raw == "":
        return ""

    direct = canonical_model(raw, candidates)
    if direct != raw:
        return direct

    norm = re.sub(r"[^a-z0-9]+", "", raw.lower())
    if norm in {"tie", "equal", "draw", "nowinner", "none", "neither", "both", "na", "n/a"}:
        return "Tie"

    answer_1_aliases = {
        "answer1", "a", "first", "modela", "answera", "optiona", "responsea", "1"
    }
    answer_2_aliases = {
        "answer2", "b", "second", "modelb", "answerb", "optionb", "responseb", "2"
    }
    if norm in answer_1_aliases:
        return answer_1_model
    if norm in answer_2_aliases:
        return answer_2_model

    # Some judges return short prose such as "Answer 1 is better".
    if re.search(r"\b(answer|model|option|response)\s*1\b|\bfirst\b", raw, flags=re.IGNORECASE):
        return answer_1_model
    if re.search(r"\b(answer|model|option|response)\s*2\b|\bsecond\b", raw, flags=re.IGNORECASE):
        return answer_2_model

    return raw

def append_winrate_row(rows, *, row, item, metric, meta, qmap, source_label, source_key):
    qid, comparison, left_model, right_model = resolve_comparison_models(row, meta, source_label)
    answer_1_model, answer_2_model, direction, is_reversed = resolve_answer_models(
        row,
        meta,
        left_model,
        right_model,
        source_label,
    )
    run_id = str(get_field(row, RUN_FIELDS, meta.get("run_id", ""))).strip()
    raw_winner = str(get_field(item, WINNER_FIELDS, "")).strip()
    if raw_winner == "":
        raise ValueError(
            f"Cannot identify winner field for {source_label}, metric={metric}. "
            f"Expected one of {list(WINNER_FIELDS)}. Available columns: {available_columns(item)}"
        )
    winner_model = normalize_winner(
        raw_winner,
        answer_1_model,
        answer_2_model,
        (left_model, right_model, answer_1_model, answer_2_model),
    )

    rows.append({
        "query_id": qid,
        "question": qmap.get(qid, ""),
        "comparison": comparison,
        "metric": metric,
        "run_id": run_id,
        "eval_run": run_id,
        "run": run_id,
        "direction": direction,
        "is_reversed": is_reversed,
        "answer_1_model": answer_1_model,
        "answer_2_model": answer_2_model,
        "raw_winner": raw_winner,
        "winner_model": winner_model,
        "left_model": left_model,
        "right_model": right_model,
        "model_a": left_model,
        "model_b": right_model,
        "winner": winner_model,
        "winner_raw": raw_winner,
        "explanation": get_field(item, EXPLANATION_FIELDS, ""),
        "source_key": source_key,
    })

def iter_winrate_items(record):
    metric = get_field(record, ("metric", "criterion", "criteria"), "")
    if metric:
        yield str(metric).strip(), record
        return

    for metric in WIN_METRICS:
        item = get_metric_item(record, metric)
        if isinstance(item, dict):
            yield metric, item

def describe_winrate_source(path, wraw):
    print("[INFO] winrate detail file:", path)
    if isinstance(wraw, dict):
        print("[INFO] winrate detail format: JSON object keyed by source_key")
        print("[INFO] source_key fields: query_id::left_model::vs::right_model::run_id::direction")
        sample_key = next(iter(wraw), "")
        sample_value = wraw.get(sample_key, {}) if sample_key else {}
        print("[INFO] sample source_key:", sample_key)
        if isinstance(sample_value, dict):
            print("[INFO] top-level detail fields:", sorted(str(k) for k in sample_value.keys()))
            for metric in WIN_METRICS:
                item = get_metric_item(sample_value, metric)
                if isinstance(item, dict):
                    print(f"[INFO] sample '{metric}' fields:", sorted(str(k) for k in item.keys()))
                    break
    elif isinstance(wraw, list):
        print("[INFO] winrate detail format: JSON list of rows")
        sample = next((r for r in wraw if isinstance(r, dict)), {})
        print("[INFO] row fields:", sorted(str(k) for k in sample.keys()))
    else:
        print("[INFO] winrate detail format:", type(wraw).__name__)

def parse_winrate(wraw, qmap):
    rows = []

    if isinstance(wraw, dict):
        records = wraw.items()
    elif isinstance(wraw, list):
        records = ((f"row_{i + 1}", row) for i, row in enumerate(wraw))
    else:
        raise ValueError(f"Unsupported winrate detail format: {type(wraw).__name__}")

    for key, val in records:
        if not isinstance(val, dict):
            continue

        meta = parse_winrate_source_key(key)
        source_label = f"source_key={key}"
        saw_metric = False
        for metric, item in iter_winrate_items(val):
            saw_metric = True
            append_winrate_row(
                rows,
                row=val,
                item=item,
                metric=metric,
                meta=meta,
                qmap=qmap,
                source_label=source_label,
                source_key=key,
            )

        if not saw_metric and isinstance(wraw, list):
            raise ValueError(
                f"Cannot identify metric field in {source_label}. "
                f"Expected a flat metric field or nested metric names. Available columns: {available_columns(val)}"
            )

    rows.sort(key=lambda r: (
        natural_qid_key(r["query_id"]),
        r["comparison"],
        r["run_id"],
        r["direction"],
        WIN_METRICS.index(r["metric"]) if r["metric"] in WIN_METRICS else 99,
    ))
    return rows

def winrate_summary(w_long):
    totals = Counter()
    wins = Counter()
    questions = {}

    for r in w_long:
        qid = r["query_id"]
        questions[qid] = r.get("question", "")
        key = (qid, r["comparison"], r["metric"])
        totals[key] += 1
        wins[(qid, r["comparison"], r["metric"], r.get("winner_model", ""))] += 1

    rows = []
    for (qid, comparison, metric), n in sorted(totals.items(), key=lambda x: (natural_qid_key(x[0][0]), x[0][1], x[0][2])):
        row = {
            "query_id": qid,
            "question": questions.get(qid, ""),
            "comparison": comparison,
            "metric": metric,
            "judgements": n,
        }
        for model in MODELS:
            c = wins[(qid, comparison, metric, model)]
            row[f"{model}_wins"] = c
            row[f"{model}_win_rate"] = round(100 * c / n, 2) if n else 0
        rows.append(row)
    return rows

def query_summary(q_wide, w_summary):
    by_q = {}

    for r in q_wide:
        qid = r["query_id"]
        row = {
            "query_id": qid,
            "question": r.get("question", ""),
        }

        # 每个模型的 Overall Score
        for model in MODELS:
            row[f"{model}_Overall"] = r.get(f"{model}_Overall Score", "")

        # EVIQUE 五项细分指标
        for metric in METRICS[:-1]:
            row[f"EVIQUE_{metric}"] = r.get(f"EVIQUE_{metric}", "")

        # EVIQUE 与每个 baseline 的 Overall 差值
        ev = row.get("EVIQUE_Overall", "")
        for model in MODELS:
            if model == "EVIQUE":
                continue
            try:
                row[f"EVIQUE_minus_{model}_Overall"] = round(float(ev) - float(row.get(f"{model}_Overall", "")), 3)
            except Exception:
                row[f"EVIQUE_minus_{model}_Overall"] = ""

        by_q[qid] = row

    # 每个 query 的 EVIQUE vs baseline pairwise Overall Winner
    for r in w_summary:
        if r.get("metric") != "Overall Winner":
            continue
        comparison = r.get("comparison", "")
        if not comparison.startswith("EVIQUE vs "):
            continue

        qid = r["query_id"]
        baseline = comparison.replace("EVIQUE vs ", "", 1)
        if qid not in by_q:
            by_q[qid] = {"query_id": qid, "question": r.get("question", "")}

        by_q[qid][f"Pairwise_EVIQUE_vs_{baseline}_EVIQUE_wins"] = r.get("EVIQUE_wins", 0)
        by_q[qid][f"Pairwise_EVIQUE_vs_{baseline}_{baseline}_wins"] = r.get(f"{baseline}_wins", 0)
        by_q[qid][f"Pairwise_EVIQUE_vs_{baseline}_EVIQUE_win_rate"] = r.get("EVIQUE_win_rate", 0)

    return [by_q[qid] for qid in sorted(by_q, key=natural_qid_key)]

def make_xlsx(path, sheets):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print("[WARN] openpyxl not available, skip xlsx:", e)
        return False

    wb = Workbook()
    wb.remove(wb.active)

    for name, rows in sheets:
        ws = wb.create_sheet(name[:31])
        if not rows:
            ws.append(["No rows parsed"])
            continue

        keys = []
        for row in rows:
            for k in row:
                if k not in keys:
                    keys.append(k)

        ws.append(keys)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(wrap_text=True)

        for row in rows:
            ws.append([row.get(k, "") for k in keys])

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for i, k in enumerate(keys, 1):
            max_len = min(max(len(k) + 2, 12), 55)
            ws.column_dimensions[get_column_letter(i)].width = max_len

    wb.save(path)
    return True

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--run-name", required=True)
    ap.add_argument("--root", default=".")
    ap.add_argument("--questions", default="")
    args = ap.parse_args()

    root = Path(args.root).resolve()
    out_root = root / "comparison_runs" / args.run_name
    eval_dir = out_root / "evaluation"
    table_dir = eval_dir / "per_query_tables_v3"
    table_dir.mkdir(parents=True, exist_ok=True)

    qmap = load_questions(args.questions)

    qraw = load_json(eval_dir / "quantitative_judgements.json")
    winrate_detail_path = eval_dir / "winrate_judgements.json"
    wraw = load_json(winrate_detail_path)
    describe_winrate_source(winrate_detail_path, wraw)

    q_long = parse_quantitative(qraw, qmap)
    q_wide = quantitative_wide(q_long)
    w_long = parse_winrate(wraw, qmap)
    w_sum = winrate_summary(w_long)
    q_sum = query_summary(q_wide, w_sum)

    write_csv(table_dir / "per_query_quantitative_long.csv", q_long)
    write_csv(table_dir / "per_query_quantitative_wide.csv", q_wide)
    write_csv(table_dir / "per_query_winrate_long.csv", w_long)
    write_csv(table_dir / "per_query_winrate_summary.csv", w_sum)
    write_csv(table_dir / "per_query_summary.csv", q_sum)

    make_xlsx(
        table_dir / "per_query_eval_tables.xlsx",
        [
            ("Query Summary", q_sum),
            ("Quant Wide", q_wide),
            ("Quant Long", q_long),
            ("Winrate Summary", w_sum),
            ("Winrate Long", w_long),
        ],
    )

    print("DONE")
    print("output dir:", table_dir)
    print("quantitative long rows:", len(q_long))
    print("quantitative wide rows:", len(q_wide))
    print("winrate long rows:", len(w_long))
    print("winrate summary rows:", len(w_sum))
    print("summary rows:", len(q_sum))
    print("xlsx:", table_dir / "per_query_eval_tables.xlsx")

if __name__ == "__main__":
    main()
